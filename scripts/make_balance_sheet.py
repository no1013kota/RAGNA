"""ゲームバランスを検討するためのExcelを作る。

数式入りなので、③パラメータの数値を書き換えると、ガチャ試算・成長試算・
デッキ試算・AIプロンプトがすべて自動で計算し直されます。

    pip install openpyxl        # 初回だけ。Bot本体には不要なライブラリです
    python scripts/make_balance_sheet.py

出力先: docs/balance/バランス設計シート.xlsx

使い魔・スキル・バランス値は ``data/master/*.json`` から読み込むため、
マスターデータを更新したあとに実行し直せば、シートも最新になります。
"""

from __future__ import annotations

import json
import sys

from itertools import combinations_with_replacement
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = Path(__file__).resolve().parents[1]
MASTER = ROOT / "data" / "master"
OUT_DIR = ROOT / "docs" / "balance"
OUT_PATH = OUT_DIR / "バランス設計シート.xlsx"

INPUT_FILL = PatternFill("solid", fgColor="FFF3C4")  # 書き換えてよい
CALC_FILL = PatternFill("solid", fgColor="EAF3FF")  # 自動計算
HEAD_FILL = PatternFill("solid", fgColor="D9D9D9")
TITLE_FILL = PatternFill("solid", fgColor="2B2D31")

TITLE_FONT = Font(bold=True, size=14, color="FFFFFF")
HEAD_FONT = Font(bold=True, size=10)
NOTE_FONT = Font(size=9, color="666666")
WARN_FONT = Font(bold=True, color="CC0000")

THIN = Side(style="thin", color="BBBBBB")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

RANKS = ("C", "B", "A", "S")
RANK_ORDER = {"S": 0, "A": 1, "B": 2, "C": 3}
GENDER = {"male": "男性", "female": "女性", "none": "なし", None: "—"}

# 各シートの名前。数式から参照するので1か所で持つ。
SH_GUIDE = "① 使い方"
SH_LIST = "② 使い魔一覧"
SH_PARAM = "③ パラメータ"
SH_GACHA = "④ ガチャ試算"
SH_GROWTH = "⑤ 成長試算"
SH_DECK = "⑥ デッキ試算"
SH_PROMPT = "⑦ AIプロンプト"

PARAM = f"'{SH_PARAM}'!"
GROWTH = f"'{SH_GROWTH}'!"

# ==================================================
# ③パラメータの行番号。数式がここを参照するので定数にしておく。
# ==================================================
R_COUNT = 5  # C=5, B=6, A=7, S=8
R_COUNT_SUM = 9
R_RATE = 13  # C=13, B=14, A=15, S=16
R_RATE_SUM = 17
R_GACHA_COST = 21
R_GROW = 25  # C=25, B=26, A=27, S=28  （B列=固定HP C列=固定ATK D列=率HP E列=率ATK）
R_MAX_LEVEL = 32
R_MATERIALS = 33
R_MODE = 34
R_SPD_VALUE = 35
R_SPD_INTERVAL = 36
R_FUSION_RATE = 40
R_SELL = 44  # C=44, B=45, A=46, S=47
R_DECK_COST = 51
R_DECK_UNITS = 52

# ⑤成長試算の行番号（ランク×3行、Lv1〜Lv10がB〜K列）
G_START = 5
G_ROWS = {rank: G_START + index * 3 for index, rank in enumerate(RANKS)}  # HP行


def load(name: str) -> dict:
    return json.loads((MASTER / f"{name}.json").read_text(encoding="utf-8"))


def title(ws, row: int, text: str, width: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = TITLE_FILL
    cell.font = TITLE_FONT
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 22


def section(ws, row: int, text: str, width: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = HEAD_FILL
    cell.font = HEAD_FONT


def note(ws, row: int, text: str, width: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
    ws.cell(row=row, column=1, value=text).font = NOTE_FONT


def header(ws, row: int, labels: list[str], start: int = 1) -> None:
    for offset, label in enumerate(labels):
        cell = ws.cell(row=row, column=start + offset, value=label)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.border = BOX
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def cell_in(ws, row: int, col: int, value, fmt: str | None = None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = INPUT_FILL
    cell.border = BOX
    if fmt:
        cell.number_format = fmt
    return cell


def cell_calc(ws, row: int, col: int, value, fmt: str | None = None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = CALC_FILL
    cell.border = BOX
    if fmt:
        cell.number_format = fmt
    return cell


def cell_flat(ws, row: int, col: int, value, fmt: str | None = None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = BOX
    if fmt:
        cell.number_format = fmt
    return cell


def widths(ws, spec: dict[str, int]) -> None:
    for column, width in spec.items():
        ws.column_dimensions[column].width = width


def grow_formula(base_ref: str, rank_row: int, level_ref: str, stat: str) -> str:
    """レベル ``level_ref`` での能力値を返す数式を組む。

    ``stat`` は "hp" か "atk"。③パラメータの「成長のしかた」を見て、
    固定値なら足し算、率なら掛け算になります。
    """

    fixed_col = "B" if stat == "hp" else "C"
    rate_col = "D" if stat == "hp" else "E"
    # ゲーム本体は 素の値 ×（1 + 成長率 × レベル）。Lv1の時点で既に1回ぶん伸びている。
    steps = f"({level_ref})"

    return (
        f'=ROUND(IF({PARAM}C{R_MODE}="固定",'
        f"{base_ref}+{PARAM}{fixed_col}{rank_row}*{steps},"
        f"{base_ref}*(1+{PARAM}{rate_col}{rank_row}*{steps})),0)"
    )


def speed_formula(base_ref: str, level_ref: str) -> str:
    """レベル ``level_ref`` でのSPDを返す数式を組む（上限あり）。"""

    bonus = f"(INT(({level_ref}-1)/{PARAM}C{R_SPD_INTERVAL})+1)*{PARAM}C{R_SPD_VALUE}"
    return f"=MIN(100,{base_ref}+{bonus})"


# ==================================================
# ① 使い方
# ==================================================
def build_guide(wb: Workbook) -> None:
    ws = wb.create_sheet(SH_GUIDE)
    widths(ws, {"A": 3, "B": 104})

    title(ws, 1, "  ラグナオンライン バランス設計シート", 2)

    lines = [
        "",
        "■ 触ってよいのは黄色のセルだけです",
        "　黄色 … あなたが書き換えるところ（③パラメータに集めてあります）",
        "　水色 … 自動計算。書き換えると数式が消えるので触らないでください",
        "",
        "■ シートの並び",
        "　② 使い魔一覧　… 全50体の能力値とスキル。Lv1とLv10を並べています",
        "　③ パラメータ　… ここだけ書き換えます。ほかのシートはここを見て計算します",
        "　④ ガチャ試算　… 最大レベルにするまでに必要な単発回数とcoin",
        "　⑤ 成長試算　　… ランクごとの平均能力値を Lv1〜Lv10 で並べたもの",
        "　⑥ デッキ試算　… 合計COST20以内の編成と、その平均ステータス",
        "　⑦ AIプロンプト … 生成AIに戦闘シミュレーションを頼むための文章",
        "",
        "■ 今回の前提",
        "　・ガチャは単発だけで計算しています（10連・保証枠は考えていません）",
        "　・プレイヤーの収入は考えていません。必要coinだけを出しています",
        "",
        "■ 成長のしかたは2通り書けます",
        "　③パラメータの「成長のしかた」を『固定』か『率』に切り替えてください。",
        "　　固定 … 合成1回ごとに決まった数だけ増える（例 HP+2）",
        "　　率　 … 素の値に対する割合で増える（例 5% ずつ）",
        "　どちらの数値も③に並べて書いてあるので、切り替えて見比べられます。",
        "",
        "■ 覚えておくと便利なこと",
        "　・最大レベルにするには「同じ使い魔」が（最大レベル−1）×素材数＋1体 必要です",
        "　・ATKの素の値は4〜11しかありません。固定で+1すると1回で10〜25%増える計算です",
        "　・SはAよりATKもSPDも低めです。Sの強みはHPとスキルとCOSTです",
        "　・スキルはATK・最大HPの割合で効くので、能力値を上げると威力も上がります",
        "",
        "■ 決まったら",
        "　数値が決まったら教えてください。data/master/*.json へ反映して",
        "　Discordへ配信します。手で書き換える必要はありません。",
        "",
        "※ このファイルは scripts/make_balance_sheet.py で作り直せます。",
        "　 能力値やスキルは data/master/ の最新の内容を読み込んでいます。",
    ]
    for index, text in enumerate(lines, start=2):
        cell = ws.cell(row=index, column=2, value=text)
        if text.startswith("■"):
            cell.font = Font(bold=True, size=11)
        elif text.startswith("※"):
            cell.font = NOTE_FONT


# ==================================================
# ② 使い魔一覧
# ==================================================
def build_list(wb: Workbook, familiars: list[dict], skills: dict) -> None:
    ws = wb.create_sheet(SH_LIST)
    widths(ws, {
        "A": 15, "B": 6, "C": 6, "D": 7,
        "E": 8, "F": 8, "G": 8,
        "H": 7, "I": 7, "J": 7,
        "K": 7, "L": 7, "M": 7,
        "N": 14, "O": 56, "P": 10,
        "Q": 14, "R": 56, "S": 16,
    })

    title(ws, 1, "  ② 使い魔一覧　— 能力値とスキル", 19)
    note(ws, 2, "　　Lv1・Lv10の能力値は③パラメータの成長設定で変わります。「素の値」とスキルは data/master/ の値です。", 19)
    note(ws, 3, "　　※ Lv1の時点で既に1回ぶん成長しています（能力値＝素の値×(1+成長率×レベル)）。", 19)

    header(ws, 4, ["", "", "", "", "素の値(Lv0)", "", "", "Lv1", "", "", "Lv10", "", "",
                   "ACTIVEスキル", "", "", "PASSIVEスキル", "", ""])
    header(ws, 5, ["名前", "ランク", "COST", "性別",
                   "HP", "ATK", "SPD", "HP", "ATK", "SPD", "HP", "ATK", "SPD",
                   "名前", "内容", "使用回数",
                   "名前", "内容", "発動タイミング"])

    trigger_label = {
        "battle_start": "バトル開始時", "on_attack": "攻撃した時",
        "on_damaged": "ダメージを受けた時", "on_defeat": "誰かが倒れた時",
        "before_defeat": "自分が倒れる直前", "turn_start": "自分のターン開始時",
        "always": "常時", "before_status_apply": "状態異常を受ける直前",
        "on_revive": "蘇生された時", "on_kill": "敵を倒した時",
        "round_start": "ラウンド開始時", "before_action": "行動の直前",
        "on_heal": "回復した時", "on_skill": "スキルを使った時",
    }

    rows = sorted(familiars, key=lambda f: (RANK_ORDER[f["rank"]], -f["base_atk"]))

    for index, fam in enumerate(rows):
        row = 6 + index
        rank_row = R_GROW + RANKS.index(fam["rank"])

        cell_flat(ws, row, 1, fam["name"])
        cell_flat(ws, row, 2, fam["rank"])
        cell_flat(ws, row, 3, fam["cost"])
        cell_flat(ws, row, 4, GENDER.get(fam.get("gender"), "—"))

        # 素の値（familiars.json の値。Lv0相当）
        cell_flat(ws, row, 5, fam["base_hp"])
        cell_flat(ws, row, 6, fam["base_atk"])
        cell_flat(ws, row, 7, fam["speed"])

        # Lv1
        cell_calc(ws, row, 8, grow_formula(f"E{row}", rank_row, "1", "hp"), "0")
        cell_calc(ws, row, 9, grow_formula(f"F{row}", rank_row, "1", "atk"), "0")
        cell_calc(ws, row, 10, speed_formula(f"G{row}", "1"), "0")

        # 最大レベル（③の設定を見る）
        level = f"{PARAM}C{R_MAX_LEVEL}"
        cell_calc(ws, row, 11, grow_formula(f"E{row}", rank_row, level, "hp"), "0")
        cell_calc(ws, row, 12, grow_formula(f"F{row}", rank_row, level, "atk"), "0")
        cell_calc(ws, row, 13, speed_formula(f"G{row}", level), "0")

        active = next(
            (skills[s] for s in fam.get("skills", [])
             if skills.get(s, {}).get("skill_type") == "active"), None)
        passive = next(
            (skills[s] for s in fam.get("skills", [])
             if skills.get(s, {}).get("skill_type") == "passive"), None)

        if active:
            cell_flat(ws, row, 14, active["name"])
            cell_flat(ws, row, 15, active["description"]).alignment = Alignment(wrap_text=True, vertical="top")
            uses = active.get("max_uses_per_battle")
            cell_flat(ws, row, 16, f"1バトル{uses}回" if uses else "制限なし")
        else:
            for col in (14, 15, 16):
                cell_flat(ws, row, col, "—")

        if passive:
            cell_flat(ws, row, 17, passive["name"])
            cell_flat(ws, row, 18, passive["description"]).alignment = Alignment(wrap_text=True, vertical="top")
            cell_flat(ws, row, 19, trigger_label.get(passive.get("trigger"), passive.get("trigger") or "—"))
        else:
            for col in (17, 18, 19):
                cell_flat(ws, row, col, "—")

    ws.freeze_panes = "E6"
    ws.auto_filter.ref = f"A5:S{5 + len(rows)}"


# ==================================================
# ③ パラメータ
# ==================================================
def build_params(wb: Workbook, balance: dict, gacha: dict, counts: dict) -> None:
    ws = wb.create_sheet(SH_PARAM)
    widths(ws, {"A": 26, "B": 13, "C": 13, "D": 13, "E": 13, "F": 50})

    pool = gacha["pools"][0]
    fam = balance["familiar"]
    battle = balance["battle"]

    title(ws, 1, "  ③ パラメータ　— 黄色のセルを書き換えてください", 6)
    note(ws, 2, "　　このシートの数値だけが計算の入り口です。ほかのシートは全部ここを見ています。", 6)

    # --- 使い魔の数 ---
    section(ws, 3, "■ 各ランクの使い魔の数", 6)
    header(ws, 4, ["ランク", "現在", "試す値", "", "", "説明"])
    for index, rank in enumerate(RANKS):
        row = R_COUNT + index
        cell_flat(ws, row, 1, f"{rank}ランク")
        cell_calc(ws, row, 2, counts[rank])
        cell_in(ws, row, 3, counts[rank])
    ws.cell(row=R_COUNT, column=6, value="数を減らすと、同じ使い魔が出やすくなります（＝合成しやすい）").font = NOTE_FONT
    cell_flat(ws, R_COUNT_SUM, 1, "合計")
    cell_calc(ws, R_COUNT_SUM, 2, f"=SUM(B{R_COUNT}:B{R_COUNT+3})")
    cell_calc(ws, R_COUNT_SUM, 3, f"=SUM(C{R_COUNT}:C{R_COUNT+3})")

    # --- 排出率 ---
    section(ws, 11, "■ ガチャの排出率（単発・千分率。合計を必ず1000にする）", 6)
    header(ws, 12, ["ランク", "現在", "試す値", "％", "", "説明"])
    for index, rank in enumerate(RANKS):
        row = R_RATE + index
        cell_flat(ws, row, 1, f"{rank}ランク")
        cell_calc(ws, row, 2, pool["rates"]["normal"][rank])
        cell_in(ws, row, 3, pool["rates"]["normal"][rank])
        cell_calc(ws, row, 4, f"=C{row}/10", '0.0"%"')
    cell_flat(ws, R_RATE_SUM, 1, "合計")
    cell_calc(ws, R_RATE_SUM, 2, f"=SUM(B{R_RATE}:B{R_RATE+3})")
    cell_calc(ws, R_RATE_SUM, 3, f"=SUM(C{R_RATE}:C{R_RATE+3})")
    ws.cell(row=R_RATE_SUM, column=6,
            value=f'=IF(C{R_RATE_SUM}=1000,"OK","★合計が1000ではありません")').font = WARN_FONT

    # --- ガチャ費用 ---
    section(ws, 19, "■ ガチャ費用", 6)
    header(ws, 20, ["項目", "現在", "試す値", "", "", "説明"])
    cell_flat(ws, R_GACHA_COST, 1, "単発の費用（coin）")
    cell_calc(ws, R_GACHA_COST, 2, pool["single_cost"], "#,##0")
    cell_in(ws, R_GACHA_COST, 3, pool["single_cost"], "#,##0")
    ws.cell(row=R_GACHA_COST, column=6, value="1回引くのにかかるcoin").font = NOTE_FONT

    # --- 成長量 ---
    section(ws, 23, "■ 合成1回あたりの成長量（固定値・％の両方を書いておけます）", 6)
    header(ws, 24, ["ランク", "固定値 HP", "固定値 ATK", "率 HP", "率 ATK", "説明"])
    fixed_default = {"C": 0, "B": 1, "A": 2, "S": 3}
    fixed_atk_default = {"C": 0, "B": 0, "A": 1, "S": 1}
    for index, rank in enumerate(RANKS):
        row = R_GROW + index
        cell_flat(ws, row, 1, f"{rank}ランク")
        cell_in(ws, row, 2, fixed_default[rank])
        cell_in(ws, row, 3, fixed_atk_default[rank])
        cell_in(ws, row, 4, fam["hp_growth_rate_per_level"])
        cell_in(ws, row, 5, fam["atk_growth_rate_per_level"])
    ws.cell(row=R_GROW, column=6,
            value="下の「成長のしかた」で、固定値と率のどちらを使うか選びます").font = NOTE_FONT
    ws.cell(row=R_GROW + 1, column=6,
            value="率は 0.05＝5%。0にするとそのランクは合成できません").font = NOTE_FONT

    # --- 成長の共通設定 ---
    section(ws, 30, "■ 成長の共通設定", 6)
    header(ws, 31, ["項目", "現在", "試す値", "", "", "説明"])
    for row, label, current, memo in (
        (R_MAX_LEVEL, "最大レベル", fam["max_level"], "下げると必要な体数が一気に減ります"),
        (R_MATERIALS, "1レベルに必要な素材数", 1, "必要な体数 =（最大レベル−1）× この数 ＋ 1"),
        (R_MODE, "成長のしかた", "率", "「固定」か「率」と入力してください"),
        (R_SPD_VALUE, "SPD増加量（1回）", fam["speed_growth_value"], "SPDは上限100で頭打ちになります"),
        (R_SPD_INTERVAL, "SPDが上がる間隔", 2, "2なら Lv1,3,5,7,9 で上がります"),
    ):
        cell_flat(ws, row, 1, label)
        cell_calc(ws, row, 2, current)
        cell_in(ws, row, 3, current)
        ws.cell(row=row, column=6, value=memo).font = NOTE_FONT

    # --- 合成費用 ---
    section(ws, 38, "■ 合成費用", 6)
    header(ws, 39, ["項目", "現在", "試す値", "", "", "説明"])
    cell_flat(ws, R_FUSION_RATE, 1, "素材1体あたり（売却額比）")
    cell_calc(ws, R_FUSION_RATE, 2, fam["fusion_cost_rate_per_material"])
    cell_in(ws, R_FUSION_RATE, 3, fam["fusion_cost_rate_per_material"])
    ws.cell(row=R_FUSION_RATE, column=6,
            value="0.5なら売却額の半分。0にすると合成は無料").font = NOTE_FONT

    # --- 売却額 ---
    section(ws, 42, "■ 売却額（Lv1）", 6)
    header(ws, 43, ["ランク", "現在", "試す値", "", "", "説明"])
    for index, rank in enumerate(RANKS):
        row = R_SELL + index
        cell_flat(ws, row, 1, f"{rank}ランク")
        cell_calc(ws, row, 2, fam["sell_base_prices"][rank], "#,##0")
        cell_in(ws, row, 3, fam["sell_base_prices"][rank], "#,##0")
    ws.cell(row=R_SELL, column=6, value="合成費用もこの金額を基準に決まります").font = NOTE_FONT

    # --- バトルの枠 ---
    section(ws, 49, "■ バトルの枠（⑥デッキ試算が使います）", 6)
    header(ws, 50, ["項目", "現在", "試す値", "", "", "説明"])
    for row, label, current, memo in (
        (R_DECK_COST, "合計COST上限", battle["max_total_cost"], "編成できる合計COST"),
        (R_DECK_UNITS, "出場できる体数", battle["max_units"], "1ギルドが出せる使い魔の数"),
    ):
        cell_flat(ws, row, 1, label)
        cell_calc(ws, row, 2, current)
        cell_in(ws, row, 3, current)
        ws.cell(row=row, column=6, value=memo).font = NOTE_FONT

    ws.freeze_panes = "A5"


# ==================================================
# ④ ガチャ試算
# ==================================================
def build_gacha(wb: Workbook) -> None:
    ws = wb.create_sheet(SH_GACHA)
    widths(ws, {"A": 18, "B": 12, "C": 16, "D": 16, "E": 16, "F": 16, "G": 16, "H": 40})

    title(ws, 1, "  ④ ガチャ試算　— 単発だけで計算しています", 8)
    note(ws, 2, "　　③パラメータを変えると自動で変わります。10連・保証枠は考慮していません。", 8)

    section(ws, 4, "■ 最大レベルに必要な体数", 8)
    cell_flat(ws, 5, 1, "必要な体数")
    cell_calc(ws, 5, 2, f"=({PARAM}C{R_MAX_LEVEL}-1)*{PARAM}C{R_MATERIALS}+1", "0")
    ws.cell(row=5, column=8, value="（最大レベル−1）× 1レベルの素材数 ＋ 本体1体").font = NOTE_FONT

    section(ws, 7, "■ ランクごとの試算", 8)
    header(ws, 8, ["ランク", "使い魔の数", "特定1体が出る確率",
                   "1体そろえる単発回数", "最大Lvに必要な単発回数",
                   "ガチャ費用", "合成費用", "合計coin"])

    for index, rank in enumerate(RANKS):
        row = 9 + index
        cnt = f"{PARAM}C{R_COUNT + index}"
        rate = f"{PARAM}C{R_RATE + index}"
        sell = f"{PARAM}C{R_SELL + index}"

        cell_flat(ws, row, 1, f"{rank}ランク")
        cell_calc(ws, row, 2, f"={cnt}")
        cell_calc(ws, row, 3, f"=IF({cnt}=0,0,{rate}/1000/{cnt})", "0.000%")
        cell_calc(ws, row, 4, f'=IF(C{row}=0,"—",1/C{row})', "#,##0")
        cell_calc(ws, row, 5, f'=IF(C{row}=0,"—",$B$5/C{row})', "#,##0")
        cell_calc(ws, row, 6, f'=IF(C{row}=0,"—",E{row}*{PARAM}C{R_GACHA_COST})', "#,##0")
        # 合成費用 = 素材の数 ×（売却額 × 費用率）
        cell_calc(ws, row, 7, f"=($B$5-1)*{sell}*{PARAM}C{R_FUSION_RATE}", "#,##0")
        cell_calc(ws, row, 8, f'=IF(C{row}=0,"—",F{row}+G{row})', "#,##0")

    section(ws, 14, "■ 参考：1体だけ引きたい場合", 8)
    header(ws, 15, ["ランク", "そのランクが出る確率", "1体出るまでの単発回数", "そのぶんのcoin", "", "", "", ""])
    for index, rank in enumerate(RANKS):
        row = 16 + index
        rate = f"{PARAM}C{R_RATE + index}"
        cell_flat(ws, row, 1, f"{rank}ランク")
        cell_calc(ws, row, 2, f"={rate}/1000", "0.0%")
        cell_calc(ws, row, 3, f'=IF({rate}=0,"—",1000/{rate})', "#,##0")
        cell_calc(ws, row, 4, f'=IF({rate}=0,"—",C{row}*{PARAM}C{R_GACHA_COST})', "#,##0")

    ws.freeze_panes = "A9"


# ==================================================
# ⑤ 成長試算
# ==================================================
def build_growth(wb: Workbook, familiars: list[dict]) -> None:
    ws = wb.create_sheet(SH_GROWTH)
    widths(ws, {"A": 10, "B": 8, "C": 11, **{chr(ord("D") + i): 8 for i in range(10)}, "N": 12})

    title(ws, 1, "  ⑤ 成長試算　— ランクごとの平均能力値を Lv1〜Lv10 で並べたもの", 14)
    note(ws, 2, "　　「素の値」はそのランクの平均です。③パラメータの成長設定を変えると全部動きます。", 14)
    note(ws, 3, "　　※ Lv1の時点で既に1回ぶん成長しています（能力値＝素の値×(1+成長率×レベル)）。", 14)

    header(
        ws, 4,
        ["ランク", "項目", "素の値(Lv0)"]
        + [f"Lv{level}" for level in range(1, 11)]
        + ["Lv1→Lv10"],
    )

    for index, rank in enumerate(RANKS):
        fams = [f for f in familiars if f["rank"] == rank]
        avg_hp = round(sum(f["base_hp"] for f in fams) / len(fams), 1)
        avg_atk = round(sum(f["base_atk"] for f in fams) / len(fams), 1)
        avg_spd = round(sum(f["speed"] for f in fams) / len(fams), 1)
        rank_row = R_GROW + index
        base = G_ROWS[rank]

        for offset, (label, value) in enumerate(
            (("HP", avg_hp), ("ATK", avg_atk), ("SPD", avg_spd))
        ):
            row = base + offset
            cell_flat(ws, row, 1, f"{rank}ランク" if offset == 0 else "")
            cell_flat(ws, row, 2, label)
            cell_flat(ws, row, 3, value, "0.0")  # 素の値（familiars.json の値＝Lv0相当）

            for level in range(1, 11):
                col = 3 + level  # 素の値がC列(3)、Lv n は (3+n) 列
                if label == "SPD":
                    formula = speed_formula("$C" + str(row), str(level))
                else:
                    formula = grow_formula(
                        "$C" + str(row), rank_row, str(level), label.lower()
                    )
                # 最大レベルを超える列は空にする
                cell_calc(
                    ws, row, col,
                    f'=IF({level}>{PARAM}C{R_MAX_LEVEL},"—",{formula[1:]})',
                    "0.0" if label != "SPD" else "0",
                )
            cell_calc(ws, row, 14, f'=IF(M{row}="—","—",M{row}/D{row})', '0.00"倍"')

    section(ws, 18, "■ 見るときのポイント", 14)
    points = [
        "　・ATKの素の値は4〜11しかありません。固定で+1すると1回で10〜25%増える計算になります。",
        "　・SはAよりATKもSPDも低めです（S:ATK8.4/SPD51 A:ATK8.6/SPD61）。SはHPとスキルとCOSTで差がついています。",
        "　・SPDは上限100で頭打ちになります。もともと速い使い魔ほど伸びしろが小さくなります。",
        "　・スキルはATK・最大HPの割合で効くため、能力値を上げるとスキル威力も一緒に上がります。",
    ]
    for offset, text in enumerate(points):
        ws.cell(row=19 + offset, column=1, value=text).font = NOTE_FONT
        ws.merge_cells(start_row=19 + offset, start_column=1, end_row=19 + offset, end_column=14)

    ws.freeze_panes = "D5"


# ==================================================
# ⑥ デッキ試算
# ==================================================
def build_deck(wb: Workbook, balance: dict) -> None:
    ws = wb.create_sheet(SH_DECK)
    widths(ws, {
        "A": 7, "B": 7, "C": 7, "D": 7, "E": 9, "F": 9,
        "G": 10, "H": 10, "I": 10, "J": 10, "K": 10, "L": 10, "M": 22,
    })

    cost = {"C": 2, "B": 3, "A": 4, "S": 5}
    max_cost = balance["battle"]["max_total_cost"]
    max_units = balance["battle"]["max_units"]

    title(ws, 1, f"  ⑥ デッキ試算　— 合計COST{max_cost}以内の編成と、その平均ステータス", 13)
    note(ws, 2, "　　ランクの組み合わせごとに、⑤成長試算の平均能力値を足し合わせています。", 13)
    note(ws, 3, "　　「合計」は5体ぶんの合計、「平均」は1体あたりです。SPDは平均で見てください。", 13)

    header(ws, 5, ["S", "A", "B", "C", "体数", "COST",
                   "Lv1 HP計", "Lv1 ATK計", "Lv1 SPD平均",
                   "Lv10 HP計", "Lv10 ATK計", "Lv10 SPD平均", "備考"])

    # 体数 max_units 以下・COST上限以内の組み合わせを列挙する
    decks = []
    for size in range(1, max_units + 1):
        for combo in combinations_with_replacement("SABC", size):
            total = sum(cost[rank] for rank in combo)
            if total <= max_cost:
                counts = {rank: combo.count(rank) for rank in "SABC"}
                decks.append((counts, size, total))

    # 満枠・高COSTを上に（強い編成から見たいので）
    decks.sort(key=lambda d: (-d[1], -d[2], -d[0]["S"], -d[0]["A"]))

    # ⑤成長試算の行（Lv1=C列、Lv10=L列）
    lv1_col, lv10_col = "D", "M"
    hp_rows = {rank: G_ROWS[rank] for rank in RANKS}
    atk_rows = {rank: G_ROWS[rank] + 1 for rank in RANKS}
    spd_rows = {rank: G_ROWS[rank] + 2 for rank in RANKS}

    def total_of(row_map: dict[str, int], col: str, deck_row: int) -> str:
        parts = [
            f"{chr(ord('A') + index)}{deck_row}*{GROWTH}{col}{row_map[rank]}"
            for index, rank in enumerate(("S", "A", "B", "C"))
        ]
        return "=" + "+".join(parts)

    for index, (counts, size, total) in enumerate(decks):
        row = 6 + index
        for offset, rank in enumerate(("S", "A", "B", "C")):
            cell_flat(ws, row, 1 + offset, counts[rank])
        cell_flat(ws, row, 5, size)
        cell_calc(ws, row, 6, total)

        cell_calc(ws, row, 7, total_of(hp_rows, lv1_col, row), "0")
        cell_calc(ws, row, 8, total_of(atk_rows, lv1_col, row), "0")
        cell_calc(ws, row, 9, total_of(spd_rows, lv1_col, row)[1:].join(["=(", f")/E{row}"]), "0")

        cell_calc(ws, row, 10, f'=IF({GROWTH}{lv10_col}{hp_rows["S"]}="—","—",{total_of(hp_rows, lv10_col, row)[1:]})', "0")
        cell_calc(ws, row, 11, f'=IF({GROWTH}{lv10_col}{atk_rows["S"]}="—","—",{total_of(atk_rows, lv10_col, row)[1:]})', "0")
        cell_calc(ws, row, 12, f'=IF({GROWTH}{lv10_col}{spd_rows["S"]}="—","—",({total_of(spd_rows, lv10_col, row)[1:]})/E{row})', "0")

        if counts["S"] == size:
            memo = "S単一"
        elif counts["C"] == size:
            memo = "C単一（最安）"
        elif total == max_cost:
            memo = "COST上限ぴったり"
        else:
            memo = ""
        cell_flat(ws, row, 13, memo)

    last = 5 + len(decks)
    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:M{last}"

    ws.cell(row=last + 2, column=1, value=f"　編成の組み合わせ: {len(decks)}通り（体数{max_units}以下・COST{max_cost}以内）").font = NOTE_FONT
    ws.merge_cells(start_row=last + 2, start_column=1, end_row=last + 2, end_column=13)
    ws.cell(row=last + 3, column=1, value="　※ ランクごとの平均値で計算しています。実際は使い魔ごとに能力値が違います。").font = NOTE_FONT
    ws.merge_cells(start_row=last + 3, start_column=1, end_row=last + 3, end_column=13)


# ==================================================
# ⑦ AIプロンプト
# ==================================================
def build_prompt(wb: Workbook) -> None:
    ws = wb.create_sheet(SH_PROMPT)
    widths(ws, {"A": 3, "B": 110})

    title(ws, 1, "  ⑦ AIプロンプト　— 生成AIに戦闘シミュレーションを頼むための文章", 2)
    note(ws, 2, "　　下のB6セルをコピーして、そのままAIに貼り付けてください。③パラメータの値が入ります。", 2)

    ws.cell(row=4, column=2, value="■ コピーする文章（③パラメータの現在値が自動で入ります）").font = Font(bold=True, size=11)

    # ③パラメータの値を埋め込んだプロンプトを数式で組み立てる
    param_line = (
        f'"【パラメータ】"&CHAR(10)&'
        f'"・使い魔の数: C="&{PARAM}C{R_COUNT}&"体 B="&{PARAM}C{R_COUNT+1}&"体 '
        f'A="&{PARAM}C{R_COUNT+2}&"体 S="&{PARAM}C{R_COUNT+3}&"体"&CHAR(10)&'
        f'"・排出率(単発): C="&{PARAM}C{R_RATE}/10&"% B="&{PARAM}C{R_RATE+1}/10&"% '
        f'A="&{PARAM}C{R_RATE+2}/10&"% S="&{PARAM}C{R_RATE+3}/10&"%"&CHAR(10)&'
        f'"・最大レベル: Lv"&{PARAM}C{R_MAX_LEVEL}&"（1レベルにつき同じ使い魔"&{PARAM}C{R_MATERIALS}&"体が必要）"&CHAR(10)&'
        f'"・成長のしかた: "&{PARAM}C{R_MODE}&CHAR(10)&'
        f'"・合成1回の成長量: C(HP"&{PARAM}B{R_GROW}&"/ATK"&{PARAM}C{R_GROW}&") '
        f'B(HP"&{PARAM}B{R_GROW+1}&"/ATK"&{PARAM}C{R_GROW+1}&") '
        f'A(HP"&{PARAM}B{R_GROW+2}&"/ATK"&{PARAM}C{R_GROW+2}&") '
        f'S(HP"&{PARAM}B{R_GROW+3}&"/ATK"&{PARAM}C{R_GROW+3}&")"&CHAR(10)&'
        f'"　（率の場合: C="&{PARAM}D{R_GROW}*100&"% B="&{PARAM}D{R_GROW+1}*100&"% '
        f'A="&{PARAM}D{R_GROW+2}*100&"% S="&{PARAM}D{R_GROW+3}*100&"%）"&CHAR(10)&'
        f'"・ガチャ単発費用: "&TEXT({PARAM}C{R_GACHA_COST},"#,##0")&" coin"&CHAR(10)&'
        f'"・合計COST上限: "&{PARAM}C{R_DECK_COST}&"／出場"&{PARAM}C{R_DECK_UNITS}&"体"'
    )

    body = (
        '"あなたはゲームバランス設計の専門家です。"&CHAR(10)&'
        '"Discord上で動く対戦ゲーム「ラグナオンライン」のバランスを見てください。"&CHAR(10)&CHAR(10)&'
        '"【ゲームの仕組み】"&CHAR(10)&'
        '"・ギルド同士が使い魔を出し合って戦うターン制バトルです。"&CHAR(10)&'
        '"・使い魔にはC/B/A/Sのランクがあり、COSTはC=2 B=3 A=4 S=5です。"&CHAR(10)&'
        '"・編成は合計COSTの上限内で組みます。SPDの高い順に行動します。"&CHAR(10)&'
        '"・同じ使い魔を合成するとレベルが上がり、HPとATKが伸びます。"&CHAR(10)&'
        '"・スキルの効果量はATKや最大HPに対する割合で決まるため、"&CHAR(10)&'
        '"　能力値を上げるとスキルの威力も一緒に上がります。"&CHAR(10)&CHAR(10)&'
        + param_line +
        '&CHAR(10)&CHAR(10)&'
        '"【お願いしたいこと】"&CHAR(10)&'
        '"1. この設定で、どのランクの使い魔が有利になるか教えてください。"&CHAR(10)&'
        '"2. 合計COST上限の中で、一番強い編成の組み方は何になりますか。"&CHAR(10)&'
        '"　 （高COSTを少数か、低COSTを多数か、どちらが有利になるか）"&CHAR(10)&'
        '"3. 壊れている（強すぎる・弱すぎる）ところがあれば指摘してください。"&CHAR(10)&'
        '"4. 直すとしたら、どのパラメータをいくつにするのが良いか提案してください。"&CHAR(10)&CHAR(10)&'
        '"【答え方】"&CHAR(10)&'
        '"・数字の根拠を必ず示してください。"&CHAR(10)&'
        '"・「なんとなく強い」ではなく、何ターンで倒せるかなどで比べてください。"&CHAR(10)&'
        '"・日本語で、専門用語を使わずに説明してください。"'
    )

    cell = ws.cell(row=6, column=2, value="=" + body)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.fill = CALC_FILL
    cell.border = BOX
    ws.row_dimensions[6].height = 420

    ws.cell(row=8, column=2, value="■ 使い魔のデータも渡したいとき").font = Font(bold=True, size=11)
    for offset, text in enumerate([
        "　②使い魔一覧のシートをコピーして、上の文章のあとに貼り付けてください。",
        "　50体ぶんの能力値とスキルが渡るので、より細かく見てもらえます。",
        "",
        "■ 結果をシートに戻すとき",
        "　AIの提案した数値を③パラメータの「試す値」列に入れると、",
        "　④ガチャ試算・⑤成長試算・⑥デッキ試算がその場で計算し直されます。",
        "　見比べて良さそうなら教えてください。data/master/*.json へ反映します。",
    ], start=9):
        cell = ws.cell(row=offset, column=2, value=text)
        if text.startswith("■"):
            cell.font = Font(bold=True, size=11)
        else:
            cell.font = NOTE_FONT


def main() -> int:
    balance = load("balance")
    gacha = load("gacha")
    familiars = load("familiars")["familiars"]
    skills = {s["skill_id"]: s for s in load("skills")["skills"]}

    counts = {rank: len([f for f in familiars if f["rank"] == rank]) for rank in RANKS}

    wb = Workbook()
    wb.remove(wb.active)

    build_guide(wb)
    build_list(wb, familiars, skills)
    build_params(wb, balance, gacha, counts)
    build_gacha(wb)
    build_growth(wb, familiars)
    build_deck(wb, balance)
    build_prompt(wb)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)

    print(f"作成しました: {OUT_PATH.relative_to(ROOT)}")
    print(f"  シート: {' / '.join(wb.sheetnames)}")
    print(f"  使い魔 {len(familiars)}体 / スキル {len(skills)}件")
    return 0


if __name__ == "__main__":
    sys.exit(main())
